import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment
import io
import os
import json

st.set_page_config(page_title="Generador de Horarios - City Market", layout="wide")

HISTORIAL_FILE = "historial_horarios.json"

def cargar_historial():
    if os.path.exists(HISTORIAL_FILE):
        try:
            with open(HISTORIAL_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            return {}
    return {}

def guardar_historial(historial):
    with open(HISTORIAL_FILE, "w", encoding="utf-8") as f:
        json.dump(historial, f, ensure_ascii=False, indent=4)

st.title("Generador de Horarios Oficial - City Market")

@st.cache_data
def cargar_catalogo():
    try:
        df = pd.read_excel("catalogo empleados.xlsx")
        df['Empleado'] = df['Empleado'].astype(str).str.strip().str.replace('.0', '', regex=False)
        df['Nombre'] = df['Nombre'].astype(str).str.strip()
        return df
    except Exception as e:
        return None

df_catalogo = cargar_catalogo()

if df_catalogo is not None:
    dept_df = df_catalogo[['Clave Departamento', 'Departamento']].drop_duplicates().reset_index(drop=True)
    lista_departamentos = dept_df['Departamento'].tolist()
    lista_nombres = sorted(df_catalogo['Nombre'].unique().tolist())
else:
    lista_departamentos = ["GERENCIA", "ABARROTES", "LACTEOS", "COCINA", "MANTENIMIENTO"]
    lista_nombres = []

hoy = datetime.today().date()
dias_hasta_miercoles = (2 - hoy.weekday()) % 7
if dias_hasta_miercoles == 0:
    dias_hasta_miercoles = 7
proximo_miercoles = hoy + timedelta(days=dias_hasta_miercoles)

opciones_semanas = []
fechas_semanas_obj = []
for i in range(4):
    f_inicio = proximo_miercoles + timedelta(days=7 * i)
    f_fin = f_inicio + timedelta(days=6)
    label = f"Miércoles {f_inicio.strftime('%d/%m/%Y')} al Martes {f_fin.strftime('%d/%m/%Y')}"
    opciones_semanas.append(label)
    fechas_semanas_obj.append((f_inicio, f_fin))

# 1. Datos Generales
st.subheader("1. Datos Generales")
col1, col2, col3, col4 = st.columns(4)

with col1:
    departamento_seleccionado = st.selectbox("Seleccione el Departamento", lista_departamentos, key="select_depto")
    no_departamento = ""
    if df_catalogo is not None:
        match_dept = dept_df[dept_df['Departamento'] == departamento_seleccionado]
        if not match_dept.empty:
            no_departamento = str(match_dept.iloc[0]['Clave Departamento'])

with col2:
    no_depto_input = st.text_input("Número de Departamento", value=no_departamento)
with col3:
    idx_semana = st.selectbox("Semana del Horario", range(len(opciones_semanas)), format_func=lambda x: opciones_semanas[x])
    f_inicio_sel, f_fin_sel = fechas_semanas_obj[idx_semana]
with col4:
    fecha_entrega = st.date_input("Fecha de Entrega", datetime.today())

historial = cargar_historial()

if 'depto_actual' not in st.session_state or st.session_state.depto_actual != departamento_seleccionado:
    st.session_state.depto_actual = departamento_seleccionado
    if departamento_seleccionado in historial:
        st.session_state.empleados = historial[departamento_seleccionado].copy()
    else:
        st.session_state.empleados = []

horarios_autorizados = [
    "Descanso",
    "Vacaciones",
    "05:00 - 13:30 (MIXTA)",
    "05:30 - 14:00 (MIXTA)",
    "06:00 - 15:00 (DIURNA)",
    "06:30 - 15:30 (DIURNA)",
    "07:00 - 16:00 (DIURNA)",
    "07:30 - 16:30 (DIURNA)",
    "08:00 - 17:00 (DIURNA)",
    "08:30 - 17:30 (DIURNA)",
    "09:00 - 18:00 (DIURNA)",
    "09:30 - 18:30 (DIURNA)",
    "10:00 - 19:00 (DIURNA)",
    "10:30 - 19:30 (DIURNA)",
    "11:00 - 20:00 (DIURNA)",
    "11:30 - 20:30 (DIURNA)",
    "12:30 - 21:00 (MIXTA)",
    "13:00 - 21:30 (MIXTA)",
    "13:30 - 22:00 (MIXTA)",
    "14:00 - 22:30 (MIXTA)",
    "14:30 - 23:00 (MIXTA)",
    "22:00 - 05:00 (NOCTURNA)",
    "22:30 - 05:30 (NOCTURNA)",
    "23:00 - 06:00 (NOCTURNA)"
]

dias_keys = ["Miércoles", "Jueves", "Viernes", "Sábado", "Domingo", "Lunes", "Martes"]

def ajustar_horario_lactancia(horario_str):
    if horario_str in ["Descanso", "Vacaciones"] or " - " not in horario_str:
        return horario_str
    try:
        partes = horario_str.split(" (")
        rango = partes[0]
        tipo = f" ({partes[1]}" if len(partes) > 1 else ""
        entrada, salida = rango.split(" - ")
        h_salida, m_salida = map(int, salida.split(":"))
        h_salida_nueva = (h_salida - 1) % 24
        salida_ajustada = f"{h_salida_nueva:02d}:{m_salida:02d}"
        return f"{entrada} - {salida_ajustada}{tipo} (LACTANCIA)"
    except:
        return horario_str

def calcular_sugerencia_comida(horario_str):
    if horario_str in ["Descanso", "Vacaciones"] or " - " not in horario_str:
        return "14:00 - 15:00"
    try:
        entrada = horario_str.split(" - ")[0]
        h, m = map(int, entrada.split(":")[:2])
        h_comida_inicio = (h + 4) % 24
        h_comida_fin = (h_comida_inicio + 1) % 24
        return f"{h_comida_inicio:02d}:{m:02d} - {h_comida_fin:02d}:{m:02d}"
    except:
        return "14:00 - 15:00"

# 2. Gestión de Empleados
st.subheader("2. Agregar o Modificar Empleado")

nombres_tabla = [emp["Nombre Completo"] for emp in st.session_state.empleados]
opciones_selector = ["-- Seleccionar colaborador --"] + list(dict.fromkeys(nombres_tabla + lista_nombres))

colaborador_seleccionado = st.selectbox("Seleccione o busque al Colaborador por Nombre", options=opciones_selector, key="select_colab_principal")

datos_precargados = None
no_emp_inicial = ""

if colaborador_seleccionado != "-- Seleccionar colaborador --":
    for emp in st.session_state.empleados:
        if emp["Nombre Completo"] == colaborador_seleccionado:
            datos_precargados = emp
            no_emp_inicial = str(emp["No. Empleado"])
            break
    
    if not no_emp_inicial and df_catalogo is not None:
        match_cat = df_catalogo[df_catalogo['Nombre'] == colaborador_seleccionado]
        if not match_cat.empty:
            no_emp_inicial = str(match_cat.iloc[0]['Empleado'])

if 'prev_colab' not in st.session_state:
    st.session_state.prev_colab = ""

if st.session_state.prev_colab != colaborador_seleccionado:
    st.session_state.prev_colab = colaborador_seleccionado
    st.session_state.input_no_emp_val = no_emp_inicial
elif 'input_no_emp_val' not in st.session_state:
    st.session_state.input_no_emp_val = no_emp_inicial

c1, c2, c3 = st.columns([2, 1, 1])

with c1:
    st.markdown(f"**Colaborador:** {colaborador_seleccionado if colaborador_seleccionado != '-- Seleccionar colaborador --' else '*(Ninguno)*'}")
with c2:
    no_empleado = st.text_input("No. de Empleado (Automático)", key="input_no_emp_val")
with c3:
    es_lactancia = st.checkbox("Hora de Lactancia (-1 hr salida)", key="chk_lactancia")

st.markdown("**Horarios Autorizados por Día (Miércoles a Martes)**")
horarios_dias = {}
cols_dias = st.columns(7)

for idx, d_key in enumerate(dias_keys):
    with cols_dias[idx]:
        default_val = "Descanso"
        if datos_precargados and d_key in datos_precargados:
            default_val = datos_precargados[d_key].replace(" (LACTANCIA)", "")
            if default_val not in horarios_autorizados:
                default_val = "Descanso"
                
        default_idx = horarios_autorizados.index(default_val) if default_val in horarios_autorizados else 0
        
        h_sel = st.selectbox(f"{d_key}", horarios_autorizados, index=default_idx, key=f"h_{d_key}_{colaborador_seleccionado}")
        if es_lactancia:
            h_sel = ajustar_horario_lactancia(h_sel)
        horarios_dias[d_key] = h_sel

comida_base = "14:00 - 15:00"
if datos_precargados and "Hora de Comida" in datos_precargados:
    comida_base = datos_precargados["Hora de Comida"]
else:
    comida_base = calcular_sugerencia_comida(horarios_dias["Miércoles"])

fc1, fc2, fc3 = st.columns(3)
with fc1:
    hora_comida_unica = st.text_input("Hora de Comida (Única para la semana)", value=comida_base, key="input_comida_unica")
with fc2:
    fecha_aviso = st.date_input("Fecha de Aviso", datetime.today(), key="date_aviso")
with fc3:
    h_aviso_val = datos_precargados["Horario de Aviso"] if datos_precargados else ""
    horario_aviso = st.text_input("Horario de Aviso", value=h_aviso_val, key="input_h_aviso")

if st.button("➕ Agregar / Actualizar Empleado en la Tabla", type="primary"):
    if colaborador_seleccionado != "-- Seleccionar colaborador --" and no_empleado:
        nuevo_emp = {
            "No. Empleado": no_empleado,
            "Nombre Completo": colaborador_seleccionado,
            "Hora de Comida": hora_comida_unica,
            "Fecha de Aviso": str(fecha_aviso),
            "Horario de Aviso": horario_aviso,
            "Semana": opciones_semanas[idx_semana]
        }
        for d_key in dias_keys:
            nuevo_emp[d_key] = horarios_dias[d_key]
            
        existente = False
        for i, emp in enumerate(st.session_state.empleados):
            if str(emp["No. Empleado"]) == str(no_empleado) or emp["Nombre Completo"] == colaborador_seleccionado:
                st.session_state.empleados[i] = nuevo_emp
                existente = True
                break
        if not existente:
            st.session_state.empleados.append(nuevo_emp)
            
        historial[departamento_seleccionado] = st.session_state.empleados
        guardar_historial(historial)
        st.success(f"Empleado {colaborador_seleccionado} registrado / actualizado correctamente.")
        st.rerun()
    else:
        st.warning("Debe seleccionar un colaborador válido y verificar su número de empleado.")

# 3. Vista Previa y Exportación
st.subheader("3. Vista Previa y Descarga del Formato Oficial")
if st.session_state.empleados:
    df_preview = pd.DataFrame(st.session_state.empleados)
    st.dataframe(df_preview, use_container_width=True)
    
    col_acc1, col_acc2 = st.columns(2)
    with col_acc1:
        if st.button("Generar y Descargar Archivo Excel"):
            template_file = "FORMATO DE HORARIO NUEVA ACTULIZACION.xlsx"
            wb = openpyxl.load_workbook(template_file)
            ws = wb['FORMATO ']
            
            ws['C4'] = departamento_seleccionado
            ws['J4'] = no_depto_input
            ws['C5'] = str(fecha_entrega)
            
            for i in range(7):
                fecha_col = f_inicio_sel + timedelta(days=i)
                ws.cell(row=7, column=4 + i, value=fecha_col)
            
            start_row = 9
            for idx, emp in enumerate(st.session_state.empleados):
                row_num = start_row + idx
                if row_num >= 23:
                    ws.insert_rows(row_num)
                
                ws.row_dimensions[row_num].height = 36
                
                ws.cell(row=row_num, column=2, value=emp["No. Empleado"])
                ws.cell(row=row_num, column=3, value=emp["Nombre Completo"])
                
                for col_idx, d_key in enumerate(dias_keys, start=4):
                    val_horario = emp[d_key]
                    cell = ws.cell(row=row_num, column=col_idx, value=val_horario)
                    
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    
                    if val_horario in ["Descanso", "Vacaciones"] or "LACTANCIA" in val_horario:
                        cell.font = Font(name="Calibri", size=16, bold=True, color="FF000000")
                    else:
                        cell.font = Font(name="Calibri", size=16, bold=False)

                ws.cell(row=row_num, column=11, value=emp["Hora de Comida"]).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.cell(row=row_num, column=12, value=emp["Fecha de Aviso"]).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=row_num, column=13, value=emp["Horario de Aviso"]).alignment = Alignment(horizontal="center", vertical="center")
            
            for col in range(4, 11):
                col_letter = openpyxl.utils.get_column_letter(col)
                ws.column_dimensions[col_letter].width = 24

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            st.download_button(
                label="📥 Descargar Horario Rellenado (Excel)",
                data=output,
                file_name=f"Horario_{departamento_seleccionado}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
    with col_acc2:
        if st.button("Limpiar Tabla de este Departamento"):
            st.session_state.empleados = []
            if departamento_seleccionado in historial:
                del historial[departamento_seleccionado]
                guardar_historial(historial)
            st.rerun()
else:
    st.info("El departamento seleccionado no tiene horarios guardados o la tabla está vacía. Agrega empleados para comenzar.")
